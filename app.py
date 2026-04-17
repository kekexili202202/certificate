from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import sqlite3
import base64
from datetime import datetime
from openpyxl import load_workbook

app = Flask(__name__, static_url_path='', static_folder='.')
CORS(app)

# 管理员密钥（可以改成你自己的密码）
ADMIN_KEY = "admin123456"

# 赛区编号映射
REGION_CODE = {
    "华北及东北赛区": "01",
    "华南赛区": "02",
    "华东赛区": "03",
    "西南赛区": "04",
    "西北赛区": "05"
}

# 反向映射（用于从编号找赛区）
CODE_TO_REGION = {v: k for k, v in REGION_CODE.items()}


# 初始化数据库
def init_db():
    conn = sqlite3.connect('data.db')
    c = conn.cursor()

    # 选手数据表（增加 region_code 和 serial_number 字段）
    c.execute('''
        CREATE TABLE IF NOT EXISTS participants (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            region TEXT NOT NULL,
            region_code TEXT NOT NULL,
            phone TEXT NOT NULL,
            organization TEXT NOT NULL,
            serial_number INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # 模板数据表
    c.execute('''
        CREATE TABLE IF NOT EXISTS templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cert_type TEXT NOT NULL,
            region TEXT NOT NULL,
            filename TEXT NOT NULL,
            pdf_data TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    conn.commit()
    conn.close()
    print("数据库初始化完成")


init_db()


def get_next_serial_number(region_code):
    """获取某个赛区的下一个序号"""
    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute('SELECT MAX(serial_number) FROM participants WHERE region_code = ?', (region_code,))
    result = c.fetchone()[0]
    conn.close()
    return (result or 0) + 1


def generate_certificate_number(region_code, serial_number, phone_last4):
    """生成证书编号 NO.NLEC2026 + 赛区编号 + 序号(4位) + 手机号后4位"""
    return f"NO.NLEC2026{region_code}{str(serial_number).zfill(4)}{phone_last4}"


# ==================== 选手数据接口 ====================

@app.route('/api/participants', methods=['GET'])
def get_participants():
    """获取所有选手数据"""
    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute('SELECT id, name, region, region_code, phone, organization, serial_number FROM participants ORDER BY id')
    rows = c.fetchall()
    conn.close()

    participants = []
    for row in rows:
        phone_last4 = row[4][-4:] if len(row[4]) >= 4 else row[4]
        cert_number = generate_certificate_number(row[3], row[6], phone_last4)
        participants.append({
            'id': row[0],
            '姓名': row[1],
            '赛区': row[2],
            '手机号': row[4],
            '所在单位': row[5],
            '证书编号': cert_number
        })
    return jsonify(participants)


@app.route('/api/participants/upload', methods=['POST'])
def upload_participants():
    """上传选手数据"""
    key = request.headers.get('X-Admin-Key')
    if key != ADMIN_KEY:
        return jsonify({'error': '无权操作'}), 401

    data = request.json
    participants = data.get('participants', [])

    conn = sqlite3.connect('data.db')
    c = conn.cursor()

    # 清空原有数据
    c.execute('DELETE FROM participants')

    # 插入新数据，自动生成序号
    for p in participants:
        region = p.get('赛区', '')
        region_code = REGION_CODE.get(region, '00')
        phone = p.get('手机号', '')
        serial_number = get_next_serial_number(region_code)

        c.execute('''
            INSERT INTO participants (name, region, region_code, phone, organization, serial_number)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (p.get('姓名', ''), region, region_code, phone, p.get('所在单位', ''), serial_number))

    conn.commit()
    conn.close()

    return jsonify({'success': True, 'count': len(participants)})


@app.route('/api/participants/query', methods=['POST'])
def query_participant():
    """查询选手信息（返回包含证书编号）"""
    data = request.json
    name = data.get('name', '').strip()
    region = data.get('region', '').strip()
    phone = data.get('phone', '').strip()
    organization = data.get('organization', '').strip()

    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute('''
        SELECT name, region, region_code, phone, organization, serial_number FROM participants
        WHERE name = ? AND region = ? AND phone = ? AND organization = ?
    ''', (name, region, phone, organization))

    row = c.fetchone()
    conn.close()

    if row:
        phone_last4 = row[3][-4:] if len(row[3]) >= 4 else row[3]
        cert_number = generate_certificate_number(row[2], row[5], phone_last4)
        return jsonify({
            'found': True,
            'participant': {
                '姓名': row[0],
                '赛区': row[1],
                '手机号': row[3],
                '所在单位': row[4],
                '证书编号': cert_number
            }
        })
    else:
        return jsonify({'found': False})


# ==================== 模板管理接口 ====================

@app.route('/api/templates', methods=['GET'])
def get_templates():
    """获取所有模板"""
    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute('SELECT cert_type, region, filename, pdf_data FROM templates')
    rows = c.fetchall()
    conn.close()

    templates = {
        'participation': {},
        'preliminary': {},
        'final': {}
    }

    for row in rows:
        cert_type = row[0]
        region = row[1]
        templates[cert_type][region] = {
            'filename': row[2],
            'pdf_data': row[3]
        }

    return jsonify(templates)


@app.route('/api/templates/upload', methods=['POST'])
def upload_template():
    """上传模板"""
    key = request.headers.get('X-Admin-Key')
    if key != ADMIN_KEY:
        return jsonify({'error': '无权操作'}), 401

    data = request.json
    cert_type = data.get('cert_type')
    region = data.get('region')
    filename = data.get('filename')
    pdf_data = data.get('pdf_data')

    conn = sqlite3.connect('data.db')
    c = conn.cursor()

    # 删除旧的同类型同赛区模板
    c.execute('DELETE FROM templates WHERE cert_type = ? AND region = ?', (cert_type, region))

    # 插入新模板
    c.execute('''
        INSERT INTO templates (cert_type, region, filename, pdf_data)
        VALUES (?, ?, ?, ?)
    ''', (cert_type, region, filename, pdf_data))

    conn.commit()
    conn.close()

    return jsonify({'success': True})


@app.route('/api/templates/delete', methods=['DELETE'])
def delete_template():
    """删除模板"""
    key = request.headers.get('X-Admin-Key')
    if key != ADMIN_KEY:
        return jsonify({'error': '无权操作'}), 401

    data = request.json
    cert_type = data.get('cert_type')
    region = data.get('region')

    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute('DELETE FROM templates WHERE cert_type = ? AND region = ?', (cert_type, region))
    conn.commit()
    conn.close()

    return jsonify({'success': True})


@app.route('/api/templates/clear', methods=['DELETE'])
def clear_templates():
    """清空所有模板"""
    key = request.headers.get('X-Admin-Key')
    if key != ADMIN_KEY:
        return jsonify({'error': '无权操作'}), 401

    conn = sqlite3.connect('data.db')
    c = conn.cursor()
    c.execute('DELETE FROM templates')
    conn.commit()
    conn.close()

    return jsonify({'success': True})


# ==================== 从服务器读取 Excel 接口 ====================

@app.route('/api/load-excel', methods=['GET'])
def load_excel():
    """从服务器读取 Excel 文件"""
    try:
        wb = load_workbook('法律英语大赛网站测试.xlsx')
        ws = wb.active

        # 获取表头（第一行）
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)

        # 找到各列的索引（第一列是赛区序号，第二列是姓名，第三列是赛区，第四列是手机号，第五列是所在学校）
        region_code_idx = headers.index('赛区序号') if '赛区序号' in headers else 0
        name_idx = headers.index('姓名') if '姓名' in headers else 1
        region_idx = headers.index('赛区') if '赛区' in headers else 2
        phone_idx = headers.index('手机号') if '手机号' in headers else 3
        org_idx = headers.index('所在学校') if '所在学校' in headers else 4

        # 读取数据（从第二行开始）
        participants = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[name_idx]:  # 如果姓名不为空
                participants.append({
                    '姓名': str(row[name_idx]) if row[name_idx] else '',
                    '赛区': str(row[region_idx]) if row[region_idx] else '',
                    '手机号': str(row[phone_idx]) if row[phone_idx] else '',
                    '所在单位': str(row[org_idx]) if row[org_idx] else ''
                })

        wb.close()
        return jsonify({'success': True, 'data': participants, 'count': len(participants)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# ==================== 页面路由 ====================

@app.route('/')
def index():
    """查询页面（分享给用户）"""
    return send_from_directory('.', 'query.html')


@app.route('/admin')
def admin():
    """管理员页面（你自己用）"""
    return send_from_directory('.', 'admin.html')


@app.route('/<path:path>')
def serve_static(path):
    """提供静态文件"""
    return send_from_directory('.', path)


if __name__ == '__main__':
    import os

    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)